import csv
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook

from .models import (
    Region, Department, Division, Section, 
    LOAStatus, ContractStatus, Employee
)
from .bulk_upload_forms import (
    RegionUploadForm, DepartmentUploadForm, DivisionUploadForm,
    SectionUploadForm, LOAStatusUploadForm,
    ContractStatusUploadForm, EmployeeUploadForm
)


# Role-based access control decorators
def is_admin_or_superuser(user):
    """Check if user is admin or superuser"""
    return user.is_superuser or user.groups.filter(name='Admin').exists()


def is_manager_or_above(user):
    """Check if user is manager, admin, or superuser"""
    return user.is_superuser or user.groups.filter(name__in=['Tender Admin', 'Tender Manager']).exists()


def is_staff_or_above(user):
    """Check if user has any tender management role"""
    return user.is_superuser or user.groups.filter(name__in=['Tender Admin', 'Tender Manager', 'Tender Staff']).exists()


@login_required
@user_passes_test(is_admin_or_superuser)
def custom_admin_dashboard(request):
    """Admin dashboard with bulk upload options"""
    context = {
        'title': 'Admin Panel',
        'lookup_models': [
            {'name': 'Region', 'url': 'tenders:bulk_upload_region', 'count': Region.objects.count()},
            {'name': 'Department', 'url': 'tenders:bulk_upload_department', 'count': Department.objects.count()},
            {'name': 'Division', 'url': 'tenders:bulk_upload_division', 'count': Division.objects.count()},
            {'name': 'Section', 'url': 'tenders:bulk_upload_section', 'count': Section.objects.count()},
            {'name': 'e-Contract Step', 'url': 'tenders:bulk_upload_loa_status', 'count': LOAStatus.objects.count()},
            {'name': 'e-Contract Status', 'url': 'tenders:bulk_upload_contract_status', 'count': ContractStatus.objects.count()},
            {'name': 'Employee', 'url': 'tenders:bulk_upload_employee', 'count': Employee.objects.count()},
        ]
    }
    return render(request, 'tenders/admin/dashboard.html', context)


def process_csv_file(file, columns):
    """Process CSV file and return list of dictionaries"""
    decoded_file = file.read().decode('utf-8-sig')  # Handle BOM
    io_string = io.StringIO(decoded_file)
    reader = csv.DictReader(io_string)
    
    # Normalize headers (strip whitespace and lowercase)
    if reader.fieldnames:
        normalized_fieldnames = {field.strip().lower(): field for field in reader.fieldnames if field}
        required_columns_lower = [col.lower() for col in columns]
        
        # Validate headers
        missing_columns = [col for col in required_columns_lower if col not in normalized_fieldnames]
        if missing_columns:
            raise ValueError(f"CSV must contain columns: {', '.join(columns)}. Found: {', '.join(reader.fieldnames)}")
        
        # Create mapping from normalized to original column names
        column_mapping = {normalized_fieldnames[col.lower()]: col for col in columns}
        
        # Process rows with normalized column names
        data = []
        for row in reader:
            normalized_row = {}
            for original_col, target_col in column_mapping.items():
                normalized_row[target_col] = row.get(original_col, '').strip() if row.get(original_col) else ''
            data.append(normalized_row)
        
        return data
    else:
        raise ValueError(f"CSV file has no headers")


def process_excel_file(file, columns):
    """Process Excel file and return list of dictionaries"""
    wb = load_workbook(file)
    ws = wb.active

    # Get headers from first row and normalize them
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    normalized_headers = {h.lower(): h for h in headers if h}
    required_columns_lower = [col.lower() for col in columns]

    # Validate headers
    missing_columns = [col for col in required_columns_lower if col not in normalized_headers]
    if missing_columns:
        raise ValueError(f"Excel file must contain columns: {', '.join(columns)}. Found: {', '.join(headers)}")

    # Create mapping from normalized to original column names
    column_mapping = {normalized_headers[col.lower()]: col for col in columns}

    # Create list of dictionaries with normalized column names
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        normalized_row = {}
        for i, header in enumerate(headers):
            if header in column_mapping and i < len(row):
                value = row[i]
                normalized_row[column_mapping[header]] = str(value).strip() if value else ''
        data.append(normalized_row)

    return data


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_region(request):
    """Bulk upload regions from CSV/Excel"""
    if request.method == 'POST':
        form = RegionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()

            try:
                # Process file based on extension
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name'])
                else:
                    data = process_excel_file(file, ['name'])

                # Create or update regions
                created_count = 0
                updated_count = 0

                for row in data:
                    if not row.get('name'):
                        continue

                    region, created = Region.objects.update_or_create(
                        name=row['name'],
                        defaults={}
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                messages.success(request, f'Successfully processed {created_count} new regions and updated {updated_count} existing regions.')
                return redirect('tenders:custom_admin_dashboard')

            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = RegionUploadForm()

    context = {
        'form': form,
        'title': 'Bulk Upload Regions',
        'model_name': 'Region',
        'required_columns': 'name',
        'example_data': 'Western Region\nEastern Region\nCentral Region'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_department(request):
    """Bulk upload departments from CSV/Excel"""
    if request.method == 'POST':
        form = DepartmentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name'])
                else:
                    data = process_excel_file(file, ['name'])
                
                created_count = 0
                updated_count = 0
                
                for row in data:
                    if not row.get('name'):
                        continue
                    
                    department, created = Department.objects.update_or_create(
                        name=row['name'],
                        defaults={}
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                
                messages.success(request, f'Successfully processed {created_count} new departments and updated {updated_count} existing departments.')
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = DepartmentUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Departments',
        'model_name': 'Department',
        'required_columns': 'name',
        'example_data': 'Human Resources\nFinance\nProcurement'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_division(request):
    """Bulk upload divisions from CSV/Excel"""
    if request.method == 'POST':
        form = DivisionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name', 'department_name'])
                else:
                    data = process_excel_file(file, ['name', 'department_name'])
                
                created_count = 0
                updated_count = 0
                errors = []
                
                for row in data:
                    if not row.get('name'):
                        continue
                    
                    try:
                        department = None
                        if row.get('department_name'):
                            department = Department.objects.get(name=row['department_name'])
                        
                        if department:
                            division, created = Division.objects.update_or_create(
                                name=row['name'],
                                department=department,
                                defaults={}
                            )
                            
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            errors.append(f"Department not specified for division '{row['name']}'")
                    except Department.DoesNotExist:
                        errors.append(f"Department '{row.get('department_name')}' not found for division '{row['name']}'")
                
                if errors:
                    for error in errors:
                        messages.warning(request, error)
                
                messages.success(request, f'Successfully processed {created_count} new divisions and updated {updated_count} existing divisions.')
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = DivisionUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Divisions',
        'model_name': 'Division',
        'required_columns': 'name, department_name',
        'example_data': 'Operations Division,Procurement\nStrategic Division,Human Resources'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_section(request):
    """Bulk upload sections from CSV/Excel"""
    if request.method == 'POST':
        form = SectionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name', 'division_name'])
                else:
                    data = process_excel_file(file, ['name', 'division_name'])
                
                created_count = 0
                updated_count = 0
                errors = []
                
                for row in data:
                    if not row.get('name'):
                        continue
                    
                    try:
                        division = None
                        if row.get('division_name'):
                            # Try to find division by name (might need department context)
                            divisions = Division.objects.filter(name=row['division_name'])
                            if divisions.count() == 1:
                                division = divisions.first()
                            elif divisions.count() > 1:
                                errors.append(f"Multiple divisions found with name '{row['division_name']}' for section '{row['name']}'. Please be more specific.")
                                continue
                            else:
                                errors.append(f"Division '{row['division_name']}' not found for section '{row['name']}'")
                                continue
                        
                        if division:
                            section, created = Section.objects.update_or_create(
                                name=row['name'],
                                division=division,
                                defaults={}
                            )
                            
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            errors.append(f"Division not specified for section '{row['name']}'")
                    except Division.DoesNotExist:
                        errors.append(f"Division '{row.get('division_name')}' not found for section '{row['name']}'")
                
                if errors:
                    for error in errors:
                        messages.warning(request, error)
                
                messages.success(request, f'Successfully processed {created_count} new sections and updated {updated_count} existing sections.')
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = SectionUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Sections',
        'model_name': 'Section',
        'required_columns': 'name, division_name',
        'example_data': 'Tender Management,Operations Division\nContract Admin,Operations Division'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_loa_status(request):
    """Bulk upload e-Contract Stepes from CSV/Excel"""
    if request.method == 'POST':
        form = LOAStatusUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name'])
                else:
                    data = process_excel_file(file, ['name'])
                
                created_count = 0
                updated_count = 0
                
                for row in data:
                    if not row.get('name'):
                        continue
                    
                    status, created = LOAStatus.objects.get_or_create(
                        name=row['name']
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                
                messages.success(request, f'Successfully processed {created_count} new e-Contract Stepes.')
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = LOAStatusUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload e-Contract Stepes',
        'model_name': 'e-Contract Step',
        'required_columns': 'name',
        'example_data': 'Pending\nApproved\nRejected'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_contract_status(request):
    """Bulk upload e-Contract Statuses from CSV/Excel"""
    if request.method == 'POST':
        form = ContractStatusUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                if file_ext == 'csv':
                    data = process_csv_file(file, ['name'])
                else:
                    data = process_excel_file(file, ['name'])
                
                created_count = 0
                updated_count = 0
                
                for row in data:
                    if not row.get('name'):
                        continue
                    
                    status, created = ContractStatus.objects.get_or_create(
                        name=row['name']
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                
                messages.success(request, f'Successfully processed {created_count} new e-Contract Statuses.')
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = ContractStatusUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload e-Contract Statuses',
        'model_name': 'e-Contract Status',
        'required_columns': 'name',
        'example_data': 'Active\nExpired\nTerminated'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
def manage_user_employee_links(request):
    """View to manage user-employee linkages"""
    from django.contrib.auth.models import User, Group
    from .models import UserProfile, Employee
    
    # Ensure all users have profiles (create missing ones)
    for user in User.objects.all():
        if not hasattr(user, 'profile'):
            # Try to find matching employee by username
            employee = Employee.objects.filter(employee_id=user.username).first()
            UserProfile.objects.create(user=user, employee=employee)
    
    # Get all users with their profiles
    users = User.objects.select_related('profile', 'profile__employee').prefetch_related('groups').all()
    
    # Get unlinked employees (no user_account)
    unlinked_employees = Employee.objects.filter(user_account__isnull=True, is_active=True).order_by(
        'last_name', 'first_name', 'employee_id'
    )
    
    # Get users without employee links
    unlinked_users = [u for u in users if not u.profile.employee]
    
    # Get all available groups
    all_groups = Group.objects.all().order_by('name')
    
    context = {
        'title': 'Manage User-Employee Links',
        'users': users,
        'unlinked_employees': unlinked_employees,
        'unlinked_users': unlinked_users,
        'all_groups': all_groups,
    }
    return render(request, 'tenders/admin/user_employee_links.html', context)


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["POST"])
def link_user_to_employee(request, user_id):
    """Link a user to an employee and assign groups"""
    from django.contrib.auth.models import User, Group
    from .models import UserProfile, Employee
    
    user = User.objects.get(pk=user_id)
    employee_id = request.POST.get('employee_id')
    group_ids = request.POST.getlist('groups')  # Get selected group IDs
    
    # Handle employee linking
    if employee_id:
        try:
            employee = Employee.objects.get(pk=employee_id)
            user.profile.employee = employee
            user.profile.save()
            messages.success(request, f'Successfully linked {user.username} to {employee.full_name}')
        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found')
    else:
        # Unlink
        user.profile.employee = None
        user.profile.save()
        messages.success(request, f'Unlinked {user.username} from employee')
    
    # Handle group assignments
    if group_ids:
        # Clear existing groups (except superuser-related groups)
        user.groups.clear()
        # Add selected groups
        for group_id in group_ids:
            try:
                group = Group.objects.get(pk=group_id)
                user.groups.add(group)
            except Group.DoesNotExist:
                pass
        messages.success(request, f'Updated group assignments for {user.username}')
    else:
        # Clear all groups if none selected
        user.groups.clear()
        messages.info(request, f'Removed all group assignments from {user.username}')
    
    return redirect('tenders:manage_user_employee_links')


@login_required
@user_passes_test(is_admin_or_superuser)
@require_http_methods(["GET", "POST"])
def bulk_upload_employee(request):
    """Bulk upload employees from CSV/Excel"""
    if request.method == 'POST':
        form = EmployeeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_ext = file.name.split('.')[-1].lower()
            
            try:
                # Required columns for employee upload
                required_columns = ['employee_id', 'first_name', 'last_name', 'email']
                
                if file_ext == 'csv':
                    data = process_csv_file(file, required_columns + ['phone', 'department_name', 'division_name', 'section_name', 'job_title', 'is_active'])
                else:
                    data = process_excel_file(file, required_columns + ['phone', 'department_name', 'division_name', 'section_name', 'job_title', 'is_active'])
                
                created_count = 0
                updated_count = 0
                skipped_count = 0
                
                for row in data:
                    # Skip rows missing required fields
                    if not all([row.get('employee_id'), row.get('first_name'), row.get('last_name'), row.get('email')]):
                        skipped_count += 1
                        continue
                    
                    # Prepare defaults dictionary
                    defaults = {
                        'first_name': row['first_name'],
                        'last_name': row['last_name'],
                        'email': row['email'],
                        'phone': row.get('phone', ''),
                        'job_title': row.get('job_title', ''),
                        'is_active': row.get('is_active', '').lower() in ['true', '1', 'yes', 'active'] if row.get('is_active') else True,
                    }
                    
                    # Handle department lookup
                    if row.get('department_name'):
                        try:
                            department = Department.objects.get(name__iexact=row['department_name'])
                            defaults['department'] = department
                        except Department.DoesNotExist:
                            pass
                    
                    # Handle division lookup
                    if row.get('division_name'):
                        try:
                            division = Division.objects.get(name__iexact=row['division_name'])
                            defaults['division'] = division
                        except Division.DoesNotExist:
                            pass
                    
                    # Handle section lookup
                    if row.get('section_name'):
                        try:
                            section = Section.objects.get(name__iexact=row['section_name'])
                            defaults['section'] = section
                        except Section.DoesNotExist:
                            pass
                    
                    # Create or update employee
                    employee, created = Employee.objects.update_or_create(
                        employee_id=row['employee_id'],
                        defaults=defaults
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                
                message = f'Successfully processed {created_count} new employees and updated {updated_count} existing employees.'
                if skipped_count > 0:
                    message += f' Skipped {skipped_count} rows with missing required fields.'
                messages.success(request, message)
                return redirect('tenders:custom_admin_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = EmployeeUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Employees',
        'model_name': 'Employee',
        'required_columns': 'employee_id, first_name, last_name, email, phone, department_name, division_name, section_name, job_title, is_active',
        'example_data': 'EMP001,John,Doe,john.doe@kengen.co.ke,0712345678,Finance,Accounts,Payroll,Accountant,true\\nEMP002,Jane,Smith,jane.smith@kengen.co.ke,0723456789,Engineering,Electrical,Generation,Engineer,true'
    }
    return render(request, 'tenders/admin/bulk_upload.html', context)
